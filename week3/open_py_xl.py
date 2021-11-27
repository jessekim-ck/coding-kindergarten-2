import openpyxl


wb = openpyxl.Workbook()
ws = wb.active
ws.title = "shit"

print(ws["A1"].value)  # None

ws["A1"] = "id"
print(ws["A1"].value)  # "id"
print(ws.cell(row=1, column=1).value)  # "id"

ws.cell(row=1, column=1, value="idx")
print(ws["A1"].value)  # "idx"

for i in range(1, 101):
    ws.cell(row=i, column=1, value=i)
    ws.cell(row=i, column=2, value=i**2)
    ws.cell(row=i, column=3, value=i**3)

wb.save("test.xlsx")